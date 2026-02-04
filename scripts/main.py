from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import json
import time
from openpyxl import load_workbook

# Load credentials from config file
import os
config_path = os.path.join(os.path.dirname(__file__), 'config.json')
with open(config_path, 'r') as f:
    config = json.load(f)

# Configuration
LOGIN_URL = "https://adtreferral.com/login/"
EMAIL = config['email']
PASSWORD = config['password']

# Setup Chrome options
chrome_options = Options()
# chrome_options.add_argument("--headless")  # Uncomment to run headless
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
# Disable Chrome password manager popups
chrome_options.add_experimental_option("prefs", {
    "credentials_enable_service": False,
    "profile.password_manager_enabled": False,
    "profile.password_manager_leak_detection": False,
})

print("Starting browser...")
driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)

try:
    # Step 1: Navigate to login page
    print(f"Step 1: Navigating to {LOGIN_URL}")
    driver.get(LOGIN_URL)
    print(f"Page loaded: {driver.title}")

    # Step 2: Find and fill login form
    print("\nStep 2: Filling login form...")

    # Wait for username field and fill it
    username_field = wait.until(EC.presence_of_element_located((By.NAME, "LoginUsername")))
    username_field.clear()
    username_field.send_keys(EMAIL)
    print(f"  ✓ Entered username: {EMAIL}")

    # Find and fill password field
    password_field = driver.find_element(By.NAME, "LoginPassword")
    password_field.clear()
    password_field.send_keys(PASSWORD)
    print(f"  ✓ Entered password: {'*' * len(PASSWORD)}")

    # Step 3: Submit the form
    print("\nStep 3: Submitting login...")

    # Find and click the login button
    # Try different selectors for the login button
    login_button = None
    button_selectors = [
        (By.CSS_SELECTOR, "button[type='submit']"),
        (By.CSS_SELECTOR, "input[type='submit']"),
        (By.CSS_SELECTOR, ".login-button"),
        (By.CSS_SELECTOR, "button.btn-primary"),
        (By.XPATH, "//button[contains(text(), 'Log')]"),
        (By.XPATH, "//button[contains(text(), 'Sign')]"),
        (By.XPATH, "//input[@value='Log In']"),
    ]

    for selector_type, selector in button_selectors:
        try:
            login_button = driver.find_element(selector_type, selector)
            if login_button.is_displayed():
                print(f"  Found login button using: {selector}")
                break
        except:
            continue

    if login_button:
        login_button.click()
        print("  ✓ Clicked login button")
    else:
        # Try submitting the form directly
        print("  No button found, submitting form via Enter key...")
        password_field.submit()

    # Step 4: Wait for response and check result
    print("\nStep 4: Waiting for login response...")
    time.sleep(3)  # Wait for AJAX response

    current_url = driver.current_url
    print(f"  Current URL: {current_url}")

    # Check if login was successful
    if 'login' not in current_url.lower() or current_url != LOGIN_URL:
        print("\n✓ LOGIN SUCCESSFUL - Redirected away from login page!")
    else:
        # Check for error messages on page
        print("\nStill on login page, checking for errors...")
        try:
            error_elements = driver.find_elements(By.CSS_SELECTOR, ".error, .alert, .validation-error, [class*='error']")
            for err in error_elements:
                if err.text.strip():
                    print(f"  Error found: {err.text.strip()}")
        except:
            pass

    # Step 5: Save cookies
    print("\n" + "="*60)
    print("Step 5: Saving session cookies...")
    print("="*60)

    cookies = driver.get_cookies()
    print(f"\nSession Cookies ({len(cookies)}):")
    cookies_dict = {}
    for cookie in cookies:
        cookies_dict[cookie['name']] = cookie['value']
        value_preview = cookie['value'][:50] + '...' if len(cookie['value']) > 50 else cookie['value']
        print(f"  {cookie['name']}: {value_preview}")

    with open('cookies.json', 'w') as f:
        json.dump(cookies_dict, f, indent=2)
    print("\n✓ Saved to cookies.json")

    # Dismiss "Check your password" popup if present
    print("\n" + "="*60)
    print("Checking for password popup...")
    print("="*60)
    time.sleep(2)  # Wait for popup to appear

    popup_dismissed = False
    # Try various ways to dismiss the popup
    close_selectors = [
        (By.CSS_SELECTOR, ".modal .close"),
        (By.CSS_SELECTOR, ".modal-close"),
        (By.CSS_SELECTOR, "button.close"),
        (By.CSS_SELECTOR, "[data-dismiss='modal']"),
        (By.CSS_SELECTOR, ".btn-close"),
        (By.XPATH, "//button[contains(@class, 'close')]"),
        (By.XPATH, "//button[contains(text(), 'Close')]"),
        (By.XPATH, "//button[contains(text(), 'OK')]"),
        (By.XPATH, "//button[contains(text(), 'Got it')]"),
        (By.XPATH, "//button[contains(text(), 'Dismiss')]"),
        (By.XPATH, "//span[contains(text(), '×')]"),
        (By.XPATH, "//*[contains(@class, 'modal')]//*[contains(@class, 'close')]"),
    ]

    for selector_type, selector in close_selectors:
        try:
            close_btn = driver.find_element(selector_type, selector)
            if close_btn.is_displayed():
                close_btn.click()
                print(f"  ✓ Dismissed popup using: {selector}")
                popup_dismissed = True
                time.sleep(1)
                break
        except:
            continue

    # If no close button found, try pressing Escape
    if not popup_dismissed:
        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            print("  ✓ Pressed Escape to dismiss popup")
            time.sleep(1)
        except:
            pass

    # Step 6: Click "Refer a Friend" button
    print("\n" + "="*60)
    print("Step 6: Looking for 'Refer a Friend' button...")
    print("="*60)

    refer_button = None
    refer_selectors = [
        (By.CSS_SELECTOR, ".EmpowermentTextLink.AllowEditorClick"),
        (By.XPATH, "//*[contains(@class, 'EmpowermentTextLink') and contains(text(), 'Refer')]"),
        (By.XPATH, "//a[contains(text(), 'Refer a Friend')]"),
        (By.XPATH, "//button[contains(text(), 'Refer a Friend')]"),
        (By.XPATH, "//*[contains(text(), 'Refer a Friend')]"),
    ]

    for selector_type, selector in refer_selectors:
        try:
            refer_button = wait.until(EC.element_to_be_clickable((selector_type, selector)))
            if refer_button.is_displayed():
                print(f"  Found 'Refer a Friend' button using: {selector}")
                break
        except:
            continue

    if refer_button:
        refer_button.click()
        print("  ✓ Clicked 'Refer a Friend' button!")
        time.sleep(3)
        print(f"  Current URL: {driver.current_url}")
        print(f"  Page title: {driver.title}")
    else:
        print("  ⚠ Could not find 'Refer a Friend' button")
        print("  Printing page source snippet to help debug...")
        links = driver.find_elements(By.TAG_NAME, "a")
        print(f"  Found {len(links)} links on page:")
        for link in links[:15]:
            text = link.text.strip()
            href = link.get_attribute("href")
            if text:
                print(f"    - '{text}' -> {href}")

    # Step 7: Read Excel and loop through all entries
    print("\n" + "="*60)
    print("Step 7: Loading Excel data...")
    print("="*60)

    xlsx_path = os.path.join(os.path.dirname(__file__), '..', 'files', 'tkinter_1.xlsx')
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Read rows (skip header row 1), carry forward business name for merged cells
    rows = []
    last_business = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        opportunity = str(row[0]).strip() if row[0] else ""
        contact = str(row[1]).strip() if row[1] else ""
        phone = str(row[2]).strip() if row[2] else ""
        if not contact or not phone:
            continue
        if opportunity:
            last_business = opportunity
        rows.append({
            'business': last_business,
            'contact': contact,
            'phone': phone,
        })

    print(f"  Found {len(rows)} valid entries to submit.\n")

    FIELD_IDS = {
        'first': 'RightArea_Suggest_amp-element-input-EmpowerSuggestFirstName-EmpowerSuggest',
        'last': 'RightArea_Suggest_amp-element-input-EmpowerSuggestLastName-EmpowerSuggest',
        'phone': 'RightArea_Suggest_amp-element-input-EmpowerSuggestPhone-EmpowerSuggest',
        'business': 'RightArea_Suggest_amp-element-input-EmpowerSuggestCompanyName-EmpowerSuggest',
    }

    SUBMIT_ID = 'RightArea_Suggest_amp-element-input-EmpowerSuggestSubmit-EmpowerSuggest'

    def fill_field(field_id, value):
        driver.execute_script(
            "var el = document.getElementById(arguments[0]);"
            "el.value = arguments[1];"
            "el.dispatchEvent(new Event('input', {bubbles: true}));"
            "el.dispatchEvent(new Event('change', {bubbles: true}));",
            field_id, value
        )

    # Skip already-submitted entries (change this number to resume from a different row)
    START_FROM = 3  # 1=KINSER, 2=MARY already done, start from 3=Jennifer

    for i, entry in enumerate(rows):
        if i < START_FROM - 1:
            continue

        business_name = entry['business']
        contact = entry['contact']
        phone = entry['phone']

        name_parts = contact.split()
        first_name = name_parts[0] if len(name_parts) >= 1 else ""
        last_name = " ".join(name_parts[1:]) if len(name_parts) >= 2 else ""

        print(f"  [{i+1}/{len(rows)}] {first_name} {last_name} | {phone} | {business_name}")

        # Wait for the first name field to be clickable (form fully ready)
        wait.until(EC.element_to_be_clickable((By.ID, FIELD_IDS['first'])))

        # Fill all fields via JS (fast)
        fill_field(FIELD_IDS['first'], first_name)
        fill_field(FIELD_IDS['last'], last_name)
        fill_field(FIELD_IDS['phone'], phone)
        fill_field(FIELD_IDS['business'], business_name)

        # Click the submit button by its ID
        submit_btn = wait.until(EC.element_to_be_clickable((By.ID, SUBMIT_ID)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", submit_btn)
        time.sleep(0.3)
        submit_btn.click()

        # Wait for submission to process
        time.sleep(2)
        print(f"    ✓ Submitted!")

        # Click the reset button to get a fresh form for the next entry
        if i < len(rows) - 1:
            reset_btn = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, ".SuggestButton.amp-element-suggest-reset-form-button")))
            reset_btn.click()
            time.sleep(1)

    print(f"\n  ✓ Done! Submitted all referrals.")

except Exception as e:
    print(f"\nError: {e}")

finally:
    print("\nBrowser will stay open. Press Enter to close it...")
    input()
    driver.quit()
    print("Browser closed.")
